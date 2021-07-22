import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
import numpy as np
import jieba
import pickle as pkl
from sklearn import metrics

UNK, PAD = '<UNK>', '<PAD>'  # 未知字，padding符号

key = {0: '军事',
       1: '体育',
       2: '汽车',
       3: '游戏',
       4: '科技',
       5: '房产',
       6: '财经',
       7: '教育',
       8: '娱乐',
       9: '其他'
       }

re_key = {'军事': 0,
          '体育': 1,
          '汽车': 2,
          '游戏': 3,
          '科技': 4,
          '房产': 5,
          '财经': 6,
          '教育': 7,
          '娱乐': 8,
          '其他': 9
          }


class Config(object):
    """配置参数"""

    def __init__(self, dataset, embedding, source=None):
        self.model_name = 'TextCNN'
        self.vocab_path = dataset + '/vocab.pkl'  # 词表
        self.weight_path = dataset + '/' + self.model_name + '.ckpt'  # 模型参数
        self.embedding_pretrained = torch.tensor(
            np.load(dataset + '/' + embedding)["embeddings"].astype('float32')) \
            if embedding != 'random' else None  # 预训练词向量
        self.text_iter = None
        self.source = source
        self.device = torch.device('cpu')
        self.dropout = 0.5  # 随机失活
        self.num_classes = 9  # 类别数
        self.n_vocab = 0  # 词表大小，在运行时赋值
        self.embed = self.embedding_pretrained.size(1) \
            if self.embedding_pretrained is not None else 300  # 字向量维度
        self.filter_sizes = (2, 3, 4)  # 卷积核尺寸
        self.num_filters = 256  # 卷积核数量(channels数)

    def build_data(self, text):
        words_line = []
        tokenizer = lambda x: jieba.lcut(x, cut_all=False)
        vocab = pkl.load(open(self.vocab_path, 'rb'))
        lin = text.strip()
        pad_size = 510
        token = tokenizer(lin)

        if len(token) < pad_size:
            token.extend([vocab.get(PAD)] * (pad_size - len(token)))
        else:
            token = token[:pad_size]
        for word in token:
            words_line.append(vocab.get(word, vocab.get(UNK)))

        return torch.tensor([words_line], dtype=torch.long)

    def load_data(self):
        tokenizer = lambda x: jieba.lcut(x, cut_all=False)
        vocab = pkl.load(open(self.vocab_path, 'rb'))
        pad_size = 510
        contents = []
        # with open(self.source, 'r', encoding='UTF-8') as f:  # txt文件
        f = openpyxl.load_workbook(self.source)
        sheets = f.get_sheet_names()
        sheet = f.get_sheet_by_name(sheets[0])
        rows = sheet.max_row
        for i in range(2, rows + 1):
            title = sheet.cell(i, 3).value
            if title is None:
                title = " "
            title = title.replace(" ", "").replace("\t", "").replace("\n", "")
            content = sheet.cell(i, 4).value
            if content is None:
                content = " "
            content = content.replace(" ", "").replace("\t", "").replace("\n", "")
            text = title + content
            lin = text.strip()
            if not lin:
                continue
            content = lin
            words_line = []
            token = tokenizer(content)
            seq_len = len(token)
            if pad_size:
                if len(token) < pad_size:
                    token.extend([PAD] * (pad_size - len(token)))
                else:
                    token = token[:pad_size]
                    seq_len = pad_size
            # word to id
            for word in token:
                words_line.append(vocab.get(word, vocab.get(UNK)))
            contents.append((words_line, 1, seq_len))
            """
            for line in f.readlines():
                lin = line.strip()
                if not lin:
                    continue
                content, label = lin.split('\t')
                words_line = []
                token = tokenizer(content)
                seq_len = len(token)
                if pad_size:
                    if len(token) < pad_size:
                        token.extend([PAD] * (pad_size - len(token)))
                    else:
                        token = token[:pad_size]
                        seq_len = pad_size
                # word to id
                for word in token:
                    words_line.append(vocab.get(word, vocab.get(UNK)))
                contents.append((words_line, int(label), seq_len))
                """
        return contents

    def build_iterator(self, source):
        iter = DatasetIterater(source)
        return iter


class DatasetIterater(object):
    def __init__(self, batches, batch_size=1, device='cpu'):
        self.batch_size = batch_size
        self.batches = batches
        self.n_batches = len(batches) // batch_size
        self.residue = False  # 记录batch数量是否为整数
        if len(batches) % self.n_batches != 0:
            self.residue = True
        self.index = 0
        self.device = device

    def _to_tensor(self, datas):
        x = torch.LongTensor([_[0] for _ in datas]).to(self.device)
        y = torch.LongTensor([_[1] for _ in datas]).to(self.device)

        # pad前的长度(超过pad_size的设为pad_size)
        seq_len = torch.LongTensor([_[2] for _ in datas]).to(self.device)
        return (x, seq_len), y

    def __next__(self):
        if self.residue and self.index == self.n_batches:
            batches = self.batches[self.index * self.batch_size: len(self.batches)]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

        elif self.index >= self.n_batches:
            self.index = 0
            raise StopIteration
        else:
            batches = self.batches[self.index * self.batch_size: (self.index + 1) * self.batch_size]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

    def __iter__(self):
        return self

    def __len__(self):
        if self.residue:
            return self.n_batches + 1
        else:
            return self.n_batches


class Model(nn.Module):
    def __init__(self, config):
        super(Model, self).__init__()
        if config.embedding_pretrained is not None:
            self.embedding = nn.Embedding.from_pretrained(config.embedding_pretrained, freeze=False)
        else:
            self.embedding = nn.Embedding(config.n_vocab, config.embed, padding_idx=config.n_vocab - 1)
        self.convs = nn.ModuleList(
            [nn.Conv2d(1, config.num_filters, (k, config.embed)) for k in config.filter_sizes])
        self.dropout = nn.Dropout(config.dropout)
        self.fc = nn.Linear(config.num_filters * len(config.filter_sizes), config.num_classes)

    def conv_and_pool(self, x, conv):
        x = F.relu(conv(x)).squeeze(3)
        x = F.max_pool1d(x, x.size(2)).squeeze(2)
        return x

    def forward(self, x):
        out = self.embedding(x[0])
        out = out.unsqueeze(1)
        out = torch.cat([self.conv_and_pool(out, conv) for conv in self.convs], 1)
        # out = self.dropout(out)
        out = self.fc(out)
        return out


def prediction_one(text):
    """输入一句话预测"""
    dataset = 'TextCNN'  # 数据集
    embedding = 'myEmbedding.npz'
    config = Config(dataset, embedding)
    model = Model(config).to(config.device)
    model.load_state_dict(torch.load(config.weight_path, map_location='cpu'))
    data = config.build_data(text)
    with torch.no_grad():
        data = data.unsqueeze(0)  # The batch dimension is missing in your input.
        outputs = model(data)
        soft_outputs = F.softmax(outputs, dim=1)
        if soft_outputs.numpy().var() < 0.01:
            num = 9  # 其他类
        else:
            num = torch.argmax(soft_outputs)
    return key[int(num)]


def prediction_batch(path):
    """批量处理文本分类"""
    dataset = 'TextCNN'  # 数据集
    embedding = 'myEmbedding.npz'
    config = Config(dataset, embedding, path)
    model = Model(config).to(config.device)
    model.load_state_dict(torch.load(config.weight_path, map_location='cpu'))
    source = config.load_data()
    text_iter = config.build_iterator(source)
    model.eval()
    predict_all = np.array([], dtype=int)
    # labels_all = np.array([], dtype=int)
    with torch.no_grad():
        for texts, labels in text_iter:
            outputs = model(texts)
            soft_outputs = F.softmax(outputs, dim=1)
            if soft_outputs.numpy().var() < 0.01:
                predic = np.array([9])  # 其他类
            else:
                predic = np.array([torch.argmax(soft_outputs)])
            # labels = labels.data.cpu().numpy()
            # labels_all = np.append(labels_all, labels)
            predict_all = np.append(predict_all, predic)
    # acc = metrics.accuracy_score(labels_all, predict_all)
    # flist = metrics.f1_score(predict_all, labels_all, average=None)
    result = []
    for idx in list(predict_all):
        result.append(key[int(idx)])
    # return acc, flist, result
    return result
