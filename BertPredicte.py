# coding: UTF-8
import jieba
import torch
import torch.nn as nn
import torch.nn.functional as F
import numpy as np
from modeling import BertModel
from tokenization import BertTokenizer
from sklearn import metrics
import openpyxl

PAD, CLS = '[PAD]', '[CLS]'  # padding符号, bert中综合信息符号

key = {0: '军事',
       1: '体育',
       2: '汽车',
       3: '游戏',
       4: '科技',
       5: '房产',
       6: '财经',
       7: '教育',
       8: '娱乐',
       9: '其他'
       }

re_key = {'军事': 0,
          '体育': 1,
          '汽车': 2,
          '游戏': 3,
          '科技': 4,
          '房产': 5,
          '财经': 6,
          '教育': 7,
          '娱乐': 8,
          '其他': 9
          }


class Config(object):
    """配置参数"""

    def __init__(self, dataset, source=None):
        self.model_name = 'bert'
        self.text_iter = None
        self.source = source
        self.class_list = ['军事', '体育', '汽车', '游戏', '科技', '房产', '财经', '教育', '娱乐']
        self.weight_path = dataset + '/' + self.model_name + '.ckpt'  # 模型训练结果
        self.device = torch.device('cpu')  # 设备

        self.num_classes = len(self.class_list)  # 类别数
        self.batch_size = 1  # mini-batch大小
        self.pad_size = 510  # 每句话处理成的长度(短填长切)
        self.bert_path = './' + dataset
        self.tokenizer = BertTokenizer.from_pretrained(self.bert_path)
        self.hidden_size = 768

    def build_data(self, text):
        pad_size = 510
        tokenizer = lambda x: jieba.lcut(x, cut_all=False)
        lin = text.strip()
        lin = tokenizer(lin)
        lin = "".join(lin)
        token = self.tokenizer.tokenize(lin)
        token = [CLS] + token
        seq_len = len(token)
        mask = []
        token_ids = self.tokenizer.convert_tokens_to_ids(token)

        if pad_size:
            if len(token) < pad_size:
                mask = [1] * len(token_ids) + [0] * (pad_size - len(token))
                token_ids += ([0] * (pad_size - len(token)))
            else:
                mask = [1] * pad_size
                token_ids = token_ids[:pad_size]
                seq_len = pad_size
        token_ids = torch.LongTensor([token_ids]).to(self.device)
        seq_len = torch.LongTensor([seq_len]).to(self.device)
        mask = torch.LongTensor([mask]).to(self.device)
        return (token_ids, seq_len, mask)

    def load_data(self):
        tokenizer = lambda x: jieba.lcut(x, cut_all=False)
        pad_size = 510
        contents = []
        # with open(self.source, 'r', encoding='UTF-8') as f:
        f = openpyxl.load_workbook(self.source)
        sheets = f.get_sheet_names()
        sheet = f.get_sheet_by_name(sheets[0])
        rows = sheet.max_row
        for i in range(2, rows+1):
            title = sheet.cell(i, 3).value
            if title is None:
                title = " "
            title = title.replace(" ", "").replace("\t", "").replace("\n", "")
            content = sheet.cell(i, 4).value
            if content is None:
                content = " "
            content = content.replace(" ", "").replace("\t", "").replace("\n", "")
            text = title + content
            lin = text.strip()
            if not lin:
                continue
            # print(lin)
            content = lin
            token = tokenizer(content)
            token = "".join(token)
            token = self.tokenizer.tokenize(token)
            token = [CLS] + token
            seq_len = len(token)
            mask = []
            token_ids = self.tokenizer.convert_tokens_to_ids(token)

            if pad_size:
                if len(token) < pad_size:
                    mask = [1] * len(token_ids) + [0] * (pad_size - len(token))
                    token_ids += ([0] * (pad_size - len(token)))
                else:
                    mask = [1] * pad_size
                    token_ids = token_ids[:pad_size]
                    seq_len = pad_size
            contents.append((token_ids, 1, seq_len, mask))
            """
            for line in f.readlines():
                lin = line.strip()
                if not lin:
                    continue
                content, label = lin.split('\t')
                token = tokenizer(content)
                token = "".join(token)
                token = self.tokenizer.tokenize(token)
                token = [CLS] + token
                seq_len = len(token)
                mask = []
                token_ids = self.tokenizer.convert_tokens_to_ids(token)

                if pad_size:
                    if len(token) < pad_size:
                        mask = [1] * len(token_ids) + [0] * (pad_size - len(token))
                        token_ids += ([0] * (pad_size - len(token)))
                    else:
                        mask = [1] * pad_size
                        token_ids = token_ids[:pad_size]
                        seq_len = pad_size
                contents.append((token_ids, int(label), seq_len, mask))
                """
        return contents

    def build_iterator(self, source):
        iter = DatasetIterater(source)
        return iter


class DatasetIterater(object):
    def __init__(self, batches, batch_size=1, device='cpu'):
        self.batch_size = batch_size
        self.batches = batches
        self.n_batches = len(batches) // batch_size
        self.residue = False  # 记录batch数量是否为整数
        if len(batches) % self.n_batches != 0:
            self.residue = True
        self.index = 0
        self.device = device

    def _to_tensor(self, datas):
        x = torch.LongTensor([_[0] for _ in datas]).to(self.device)
        y = torch.LongTensor([_[1] for _ in datas]).to(self.device)

        # pad前的长度(超过pad_size的设为pad_size)
        seq_len = torch.LongTensor([_[2] for _ in datas]).to(self.device)
        mask = torch.LongTensor([_[3] for _ in datas]).to(self.device)
        return (x, seq_len, mask), y

    def __next__(self):
        if self.residue and self.index == self.n_batches:
            batches = self.batches[self.index * self.batch_size: len(self.batches)]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

        elif self.index >= self.n_batches:
            self.index = 0
            raise StopIteration
        else:
            batches = self.batches[self.index * self.batch_size: (self.index + 1) * self.batch_size]
            self.index += 1
            batches = self._to_tensor(batches)
            return batches

    def __iter__(self):
        return self

    def __len__(self):
        if self.residue:
            return self.n_batches + 1
        else:
            return self.n_batches


class Model(nn.Module):

    def __init__(self, config):
        super(Model, self).__init__()
        self.bert = BertModel.from_pretrained(config.bert_path)
        for param in self.bert.parameters():
            param.requires_grad = True
        self.fc = nn.Linear(config.hidden_size, config.num_classes)

    def forward(self, x):
        context = x[0]  # 输入的句子
        mask = x[2]  # 对padding部分进行mask，和句子一个size，padding部分用0表示，如：[1, 1, 1, 1, 0, 0]
        _, pooled = self.bert(context, attention_mask=mask, output_all_encoded_layers=False)
        out = self.fc(pooled)
        return out


def prediction_one(text):
    """输入一句话预测"""
    dataset = 'Bert'  # 数据集
    config = Config(dataset)
    model = Model(config).to(config.device)
    model.load_state_dict(torch.load(config.weight_path, map_location='cpu'))
    data = config.build_data(text)
    with torch.no_grad():
        model.eval()
        outputs = model(data)
        soft_outputs = F.softmax(outputs, dim=1)
        if soft_outputs.numpy().var() < 0.01:
            num = 9  # 其他类
        else:
            num = torch.argmax(soft_outputs)
    return key[int(num)]


def prediction_batch(path):
    """批量处理文本分类"""
    dataset = 'Bert'  # 数据集
    config = Config(dataset, path)
    model = Model(config).to(config.device)
    model.load_state_dict(torch.load(config.weight_path, map_location='cpu'))
    source = config.load_data()
    text_iter = config.build_iterator(source)
    model.eval()
    predict_all = np.array([], dtype=int)
    labels_all = np.array([], dtype=int)
    with torch.no_grad():
        for texts, labels in text_iter:
            outputs = model(texts)
            soft_outputs = F.softmax(outputs, dim=1)
            if soft_outputs.numpy().var() < 0.01:
                predic = np.array([9])  # 其他类
            else:
                predic = np.array([torch.argmax(soft_outputs)])
            # labels = labels.data.cpu().numpy()
            # labels_all = np.append(labels_all, labels)
            predict_all = np.append(predict_all, predic)
    # acc = metrics.accuracy_score(labels_all, predict_all)
    # flist = metrics.f1_score(predict_all, labels_all, average=None)
    result = []
    for idx in list(predict_all):
        result.append(key[int(idx)])
    # return acc, flist, result
    return result
