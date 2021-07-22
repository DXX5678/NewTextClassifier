import os

import openpyxl
from PyQt5.QtCore import pyqtSlot
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMainWindow, QMessageBox

import BertPredicte
import CNNPredicte
from MainUI import Ui_MainWindow


class TextSource:
    def __init__(self, UserInterface):
        if UserInterface.titlecontent.text():
            if UserInterface.contentcontent.toPlainText():
                self.text = UserInterface.titlecontent.text() + UserInterface.contentcontent.toPlainText()
            else:
                self.text = UserInterface.titlecontent.text()
        else:
            if UserInterface.contentcontent.toPlainText():
                self.text = UserInterface.contentcontent.toPlainText()
            else:
                self.text = None


class BatchTextSource:
    def __init__(self, UserInterface):
        self.texts = UserInterface.path


class UserInterface(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(UserInterface, self).__init__(parent)
        self.setupUi(self)
        self.retranslateUi(self)

    @pyqtSlot()
    def on_clearing_clicked(self):
        if QMessageBox.question(self, "提示", "确认清空全部内容?", QMessageBox.Yes | QMessageBox.No,
                                QMessageBox.No) == QMessageBox.Yes:
            self.contentcontent.clear()
            self.titlecontent.clear()
            self.classification.clear()
            self.fscore.clear()

    @pyqtSlot()
    def on_action_clicked(self):
        self.text = TextSource(self).text
        if self.text is None:
            QMessageBox.information(self, "提示", "请输入文本！", QMessageBox.Ok)
        else:
            # result = CNNPredicte.prediction_one(self.text)
            result = BertPredicte.prediction_one(self.text)
            self.classification.setText(result)

    @pyqtSlot()
    def on_batchaction_clicked(self):
        fileopenwindow = QtWidgets.QMainWindow()
        pathtemp = QtWidgets.QFileDialog.getOpenFileName(fileopenwindow, "选择文件", os.getcwd(), "Text Files ("
                                                                                              "*.xlsx);;All Files (*)")
        if pathtemp[0]:
            self.path = pathtemp[0]
            self.texts = BatchTextSource(self).texts
            # acc, report, result = CNNPredicte.prediction_batch(self.texts)
            # acc, report, result = BertPredicte.prediction_batch(self.texts)
            # result = CNNPredicte.prediction_batch(self.texts)
            result = BertPredicte.prediction_batch(self.texts)
            """
            self.fscore.setText("准确率:" + str(acc))
            if report.size == 10:
                self.fscore.append("军事（F1_score）:" + str(report[0]))
                self.fscore.append("体育（F1_score）:" + str(report[1]))
                self.fscore.append("汽车（F1_score）:" + str(report[2]))
                self.fscore.append("游戏（F1_score）:" + str(report[3]))
                self.fscore.append("科技（F1_score）:" + str(report[4]))
                self.fscore.append("房产（F1_score）:" + str(report[5]))
                self.fscore.append("财经（F1_score）:" + str(report[6]))
                self.fscore.append("教育（F1_score）:" + str(report[7]))
                self.fscore.append("娱乐（F1_score）:" + str(report[8]))
                self.fscore.append("其他（F1_score）:" + str(report[9]))
            else:
                self.fscore.append("军事（F1_score）:" + str(report[0]))
                self.fscore.append("体育（F1_score）:" + str(report[1]))
                self.fscore.append("汽车（F1_score）:" + str(report[2]))
                self.fscore.append("游戏（F1_score）:" + str(report[3]))
                self.fscore.append("科技（F1_score）:" + str(report[4]))
                self.fscore.append("房产（F1_score）:" + str(report[5]))
                self.fscore.append("财经（F1_score）:" + str(report[6]))
                self.fscore.append("教育（F1_score）:" + str(report[7]))
                self.fscore.append("娱乐（F1_score）:" + str(report[8]))
            """
            f = openpyxl.load_workbook(self.texts)
            sheets = f.get_sheet_names()
            sheet = f.get_sheet_by_name(sheets[0])
            sheet.cell(2, 2).value = result[0]
            self.classification.setText(result[0])
            for i in range(1, len(result)):
                sheet.cell(i+2, 2).value = result[i]
                self.classification.append(result[i])
            f.save(self.texts)
        else:
            QMessageBox.information(self, "提示", "读取文件失败！", QMessageBox.Ok)
